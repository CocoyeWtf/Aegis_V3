"""
Service d'import CSV/Excel / CSV/Excel import service.
Parse les fichiers et retourne des listes de dictionnaires.
"""

import csv
import io
import math
from collections import defaultdict
from typing import Any

import xlrd
from openpyxl import load_workbook


class ImportService:
    """Import de données depuis fichiers / Data import from files."""

    @staticmethod
    def parse_csv(content: bytes) -> list[dict[str, Any]]:
        """Parser un fichier CSV / Parse a CSV file."""
        text = content.decode("utf-8-sig")  # BOM-safe
        reader = csv.DictReader(io.StringIO(text), delimiter=";")
        # Essayer aussi avec la virgule / Try comma delimiter too
        rows = list(reader)
        if not rows or len(rows[0]) <= 1:
            reader = csv.DictReader(io.StringIO(text), delimiter=",")
            rows = list(reader)
        return rows

    @staticmethod
    def parse_excel(content: bytes) -> list[dict[str, Any]]:
        """Parser un fichier Excel (.xlsx ou .xls) / Parse an Excel file (.xlsx or .xls)."""
        if ImportService._is_xls(content):
            return ImportService._parse_excel_xls(content)

        wb = load_workbook(filename=io.BytesIO(content), read_only=True)
        ws = wb.active
        if ws is None:
            return []

        rows_iter = ws.iter_rows(values_only=True)
        headers = next(rows_iter, None)
        if not headers:
            return []

        # Nettoyer les en-têtes / Clean headers
        # Alias de colonnes connus / Known column aliases
        _COL_ALIASES = {
            "distance": "distance_km",
        }
        clean_headers = []
        for i, h in enumerate(headers):
            if h:
                key = str(h).strip().lower().replace(" ", "_")
                key = _COL_ALIASES.get(key, key)
            else:
                key = f"col_{i}"
            clean_headers.append(key)

        result = []
        for row in rows_iter:
            record = {}
            for key, val in zip(clean_headers, row):
                record[key] = val
            if any(v is not None for v in record.values()):
                result.append(record)

        wb.close()
        return result

    @staticmethod
    def _parse_excel_xls(content: bytes) -> list[dict[str, Any]]:
        """Parser un fichier .xls via xlrd / Parse a .xls file via xlrd."""
        wb = xlrd.open_workbook(file_contents=content)
        ws = wb.sheet_by_index(0)
        if ws.nrows < 2:
            return []

        _COL_ALIASES = {
            "distance": "distance_km",
        }
        clean_headers = []
        for c in range(ws.ncols):
            h = ws.cell_value(0, c)
            if h:
                key = str(h).strip().lower().replace(" ", "_")
                key = _COL_ALIASES.get(key, key)
            else:
                key = f"col_{c}"
            clean_headers.append(key)

        result = []
        for r in range(1, ws.nrows):
            record = {}
            for c in range(ws.ncols):
                key = clean_headers[c] if c < len(clean_headers) else f"col_{c}"
                record[key] = ws.cell_value(r, c)
            if any(v is not None and v != "" for v in record.values()):
                result.append(record)

        return result

    @staticmethod
    def _is_xls(content: bytes) -> bool:
        """Détecter le format .xls (BIFF) vs .xlsx (ZIP) / Detect .xls vs .xlsx format."""
        # Les fichiers .xls commencent par la signature OLE2 (D0 CF 11 E0)
        # Les fichiers .xlsx commencent par PK (50 4B) — c'est un ZIP
        return content[:4] == b"\xd0\xcf\x11\xe0" or not content[:2] == b"PK"

    @staticmethod
    def is_superlog(content: bytes) -> bool:
        """Détecter le format SUPERLOG / Detect SUPERLOG format.
        Critères : cellule A1 contient "Attendre" OU ligne 5 contient "Lieu final de livraison".
        """
        try:
            if ImportService._is_xls(content):
                return ImportService._is_superlog_xls(content)
            wb = load_workbook(filename=io.BytesIO(content), read_only=True)
            ws = wb["in"] if "in" in wb.sheetnames else wb.active
            if ws is None:
                wb.close()
                return False
            a1 = None
            row5_vals = []
            for i, row in enumerate(ws.iter_rows(max_row=5, values_only=True), 1):
                if i == 1:
                    a1 = row[0] if row else None
                if i == 5:
                    row5_vals = list(row)
            wb.close()
            if a1 and "attendre" in str(a1).lower():
                return True
            for v in row5_vals:
                if v and "lieu final" in str(v).lower():
                    return True
        except Exception:
            pass
        return False

    @staticmethod
    def _is_superlog_xls(content: bytes) -> bool:
        """Détecter SUPERLOG dans un .xls / Detect SUPERLOG in .xls file."""
        try:
            wb = xlrd.open_workbook(file_contents=content)
            ws = wb.sheet_by_name("in") if "in" in wb.sheet_names() else wb.sheet_by_index(0)
            a1 = ws.cell_value(0, 0) if ws.nrows > 0 else None
            if a1 and "attendre" in str(a1).lower():
                return True
            if ws.nrows >= 5:
                for c in range(ws.ncols):
                    v = ws.cell_value(4, c)
                    if v and "lieu final" in str(v).lower():
                        return True
        except Exception:
            pass
        return False

    # Mapping colonnes SUPERLOG → noms internes / SUPERLOG column mapping
    _SUPERLOG_COL_MAP = {
        "lieu final de livraison": "pdv_id",
        "colis": "nb_colis",
        "poids brut (kg)": "weight_kg",
        "volume (m3)": "volume_m3",
        "eqc": "_eqc",
        "eqp": "_eqp",
    }

    @staticmethod
    def _aggregate_superlog(headers: list[str | None], rows: list[list]) -> list[dict[str, Any]]:
        """Agréger des lignes SUPERLOG par PDV / Aggregate SUPERLOG rows by PDV."""
        agg: dict[str, dict[str, float]] = defaultdict(lambda: {
            "nb_colis": 0.0, "weight_kg": 0.0, "volume_m3": 0.0,
            "_eqc": 0.0, "nb_supports": 0,
        })

        for row in rows:
            pdv_code = None
            row_data: dict[str, float] = {}
            for col_idx, val in enumerate(row):
                if col_idx >= len(headers) or headers[col_idx] is None:
                    continue
                field = headers[col_idx]
                if field == "pdv_id":
                    if val is not None:
                        pdv_code = str(int(val)) if isinstance(val, (int, float)) else str(val).strip()
                else:
                    try:
                        row_data[field] = float(val) if val is not None else 0.0
                    except (ValueError, TypeError):
                        row_data[field] = 0.0

            if not pdv_code:
                continue

            bucket = agg[pdv_code]
            bucket["nb_colis"] += row_data.get("nb_colis", 0.0)
            bucket["weight_kg"] += row_data.get("weight_kg", 0.0)
            bucket["volume_m3"] += row_data.get("volume_m3", 0.0)
            bucket["_eqc"] += row_data.get("_eqc", 0.0)
            bucket["nb_supports"] += 1

        result: list[dict[str, Any]] = []
        for pdv_code, data in agg.items():
            result.append({
                "pdv_id": pdv_code,
                "nb_colis": int(data["nb_colis"]),
                "weight_kg": round(data["weight_kg"], 2),
                "volume_m3": round(data["volume_m3"], 4),
                "eqp_count": math.ceil(data["_eqc"]),
                "nb_supports": int(data["nb_supports"]),
            })
        return result

    @staticmethod
    def parse_superlog_excel(content: bytes) -> list[dict[str, Any]]:
        """Parser un fichier SUPERLOG et agréger par PDV / Parse SUPERLOG file and aggregate by PDV.
        Headers en ligne 5, données ligne 6+. Agrège par pdv_id : sum(colis, weight, volume_m3, EQP), count(supports).
        Supporte .xlsx (openpyxl) et .xls (xlrd).
        """
        if ImportService._is_xls(content):
            return ImportService._parse_superlog_xls(content)

        wb = load_workbook(filename=io.BytesIO(content), read_only=True)
        ws = wb["in"] if "in" in wb.sheetnames else wb.active
        if ws is None:
            wb.close()
            return []

        col_map = ImportService._SUPERLOG_COL_MAP

        # Lire headers ligne 5 / Read headers at row 5
        headers: list[str | None] = []
        for i, row in enumerate(ws.iter_rows(max_row=5, values_only=True), 1):
            if i == 5:
                for h in row:
                    if h:
                        key = str(h).strip().lower()
                        headers.append(col_map.get(key))
                    else:
                        headers.append(None)

        # Lire toutes les lignes de données / Read all data rows
        data_rows = [list(row) for row in ws.iter_rows(min_row=6, values_only=True)]
        wb.close()

        return ImportService._aggregate_superlog(headers, data_rows)

    @staticmethod
    def _parse_superlog_xls(content: bytes) -> list[dict[str, Any]]:
        """Parser un fichier SUPERLOG .xls via xlrd / Parse SUPERLOG .xls file via xlrd."""
        wb = xlrd.open_workbook(file_contents=content)
        ws = wb.sheet_by_name("in") if "in" in wb.sheet_names() else wb.sheet_by_index(0)

        col_map = ImportService._SUPERLOG_COL_MAP

        # Headers ligne 5 (index 4) / Headers at row 5 (index 4)
        headers: list[str | None] = []
        if ws.nrows >= 5:
            for c in range(ws.ncols):
                h = ws.cell_value(4, c)
                if h:
                    key = str(h).strip().lower()
                    headers.append(col_map.get(key))
                else:
                    headers.append(None)

        # Données ligne 6+ (index 5+) / Data from row 6+ (index 5+)
        data_rows = []
        for r in range(5, ws.nrows):
            data_rows.append([ws.cell_value(r, c) for c in range(ws.ncols)])

        return ImportService._aggregate_superlog(headers, data_rows)

    @staticmethod
    def parse_file(content: bytes, filename: str) -> list[dict[str, Any]]:
        """Parser un fichier selon son extension / Parse file based on extension."""
        ext = filename.rsplit(".", 1)[-1].lower()
        if ext == "csv":
            return ImportService.parse_csv(content)
        elif ext in ("xlsx", "xls"):
            # Détecter le format SUPERLOG avant le parse standard / Detect SUPERLOG before standard parse
            if ImportService.is_superlog(content):
                return ImportService.parse_superlog_excel(content)
            return ImportService.parse_excel(content)
        raise ValueError(f"Unsupported file type: {ext}")

    # Mapping des champs attendus par entité / Expected field mapping per entity
    ENTITY_FIELDS: dict[str, list[str]] = {
        "countries": ["name", "code"],
        "regions": ["name", "country_id"],
        "bases": ["code", "name", "type", "address", "postal_code", "city", "phone", "email", "latitude", "longitude", "region_id"],
        "pdvs": ["code", "name", "type", "address", "postal_code", "city", "phone", "email", "latitude", "longitude",
                 "has_sas", "sas_capacity", "has_dock", "dock_time_minutes", "unload_time_per_eqp_minutes",
                 "delivery_window_start", "delivery_window_end", "access_constraints", "allowed_vehicle_types", "region_id"],
        "suppliers": ["code", "name", "address", "postal_code", "city", "phone", "email", "latitude", "longitude", "region_id"],
        "volumes": ["pdv_id", "date", "nb_colis", "eqp_count", "weight_kg", "temperature_class", "base_origin_id", "preparation_start", "preparation_end", "dispatch_date", "dispatch_time", "activity_type", "promo_start_date", "volume_m3", "nb_supports"],
        "contracts": ["code", "transporter_name", "fixed_daily_cost", "vacation", "cost_per_km", "cost_per_hour",
                       "min_hours_per_day", "min_km_per_day", "consumption_coefficient",
                       "start_date", "end_date", "region_id",
                       "vehicle_code", "vehicle_name", "temperature_type", "vehicle_type",
                       "capacity_eqp", "capacity_weight_kg", "has_tailgate", "tailgate_type"],
        "distances": ["origin_type", "origin_id", "destination_type", "destination_id", "distance_km", "duration_minutes"],
        "km-tax": ["origin_type", "origin_id", "destination_type", "destination_id", "tax_per_km"],
    }
