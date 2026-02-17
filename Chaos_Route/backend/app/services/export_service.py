"""
Service d'export CSV/Excel / CSV/Excel export service.
Génère des fichiers CSV et XLSX à partir de listes de dictionnaires.
"""

import csv
import io
from typing import Any

from openpyxl import Workbook

from app.services.import_service import ImportService


class ExportService:
    """Export de données vers CSV/XLSX / Data export to CSV/XLSX."""

    @staticmethod
    def model_to_dict(obj: Any, fields: list[str]) -> dict[str, Any]:
        """Extraire les attributs d'un modèle SQLAlchemy / Extract model attributes to dict."""
        result = {}
        for f in fields:
            val = getattr(obj, f, None)
            if val is True:
                val = "true"
            elif val is False:
                val = "false"
            result[f] = val
        return result

    @staticmethod
    def to_csv(rows: list[dict], fields: list[str]) -> bytes:
        """Générer un CSV UTF-8 BOM avec séparateur ';' / Generate UTF-8 BOM CSV with ';' separator."""
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=fields, delimiter=";", extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({f: row.get(f, "") for f in fields})
        return ("\ufeff" + output.getvalue()).encode("utf-8")

    @staticmethod
    def to_xlsx(rows: list[dict], fields: list[str], sheet_name: str = "Data") -> bytes:
        """Générer un fichier Excel / Generate an Excel file."""
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        # En-têtes / Headers
        for col_idx, field in enumerate(fields, 1):
            cell = ws.cell(row=1, column=col_idx, value=field)
            cell.font = cell.font.copy(bold=True)

        # Données / Data rows
        for row_idx, row in enumerate(rows, 2):
            for col_idx, field in enumerate(fields, 1):
                ws.cell(row=row_idx, column=col_idx, value=row.get(field))

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    @staticmethod
    def get_fields(entity_type: str) -> list[str]:
        """Récupérer les champs pour une entité / Get fields for an entity type."""
        return ImportService.ENTITY_FIELDS.get(entity_type, [])
