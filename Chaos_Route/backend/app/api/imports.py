"""Routes Import CSV/Excel / Import API routes."""

from datetime import time as dt_time

from fastapi import APIRouter, Depends, UploadFile, File, HTTPException, Query
from sqlalchemy import select, inspect, func, delete as sa_delete
from sqlalchemy.ext.asyncio import AsyncSession

from app.database import get_db
from app.models.country import Country
from app.models.region import Region
from app.models.base_logistics import BaseLogistics
from app.models.pdv import PDV
from app.models.supplier import Supplier
from app.models.volume import Volume
from app.models.contract import Contract
from app.models.distance_matrix import DistanceMatrix
from app.models.km_tax import KmTax
from app.models.user import User
from app.services.import_service import ImportService
from app.api.deps import require_permission

router = APIRouter()

# Mapping entité -> modèle SQLAlchemy / Entity to model mapping
ENTITY_MODEL_MAP = {
    "countries": Country,
    "regions": Region,
    "bases": BaseLogistics,
    "pdvs": PDV,
    "suppliers": Supplier,
    "volumes": Volume,
    "contracts": Contract,
    "distances": DistanceMatrix,
    "km-tax": KmTax,
}

# Champs obligatoires par entité / Required fields per entity
REQUIRED_FIELDS = {
    "countries": {"name", "code"},
    "regions": {"name", "country_id"},
    "bases": {"code", "name", "region_id"},
    "pdvs": {"code", "name", "type", "region_id"},
    "suppliers": {"code", "name", "region_id"},
    "volumes": {"pdv_id", "date", "eqp_count", "base_origin_id"},
    "contracts": {"code", "transporter_name", "region_id"},
    "distances": {"origin_type", "origin_id", "destination_type", "destination_id", "distance_km", "duration_minutes"},
    "km-tax": {"origin_type", "origin_id", "destination_type", "destination_id", "tax_per_km"},
}

_NA_VALUES = {"#n/a", "#na", "n/a", "na", "#ref!", "#value!", "#div/0!", "-", "null", "none", "nan"}

_PDV_TYPE_ALIASES = {
    "express": "EXPRESS",
    "contact": "CONTACT",
    "super": "SUPER_ALIMENTAIRE",
    "super alimentaire": "SUPER_ALIMENTAIRE",
    "super_alimentaire": "SUPER_ALIMENTAIRE",
    "super generaliste": "SUPER_GENERALISTE",
    "super_generaliste": "SUPER_GENERALISTE",
    "hyper": "HYPER",
    "hypermarche": "HYPER",
    "hypermarché": "HYPER",
    "netto": "NETTO",
    "drive": "DRIVE",
    "urbain proxi": "URBAIN_PROXI",
    "urbain_proxi": "URBAIN_PROXI",
    "proxi": "URBAIN_PROXI",
}


def _coerce_value(val, field_name: str):
    """Convertir les valeurs selon le nom du champ / Coerce values based on field name."""
    if val is None or val == "":
        return None

    # Gérer les types natifs openpyxl avant str() / Handle native openpyxl types before str()
    from datetime import datetime as dt_datetime, date as dt_date
    if isinstance(val, dt_datetime):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, dt_date):
        return val.isoformat()
    if isinstance(val, dt_time):
        return val.strftime("%H:%M")

    s = str(val).strip()
    if s == "" or s.lower() in _NA_VALUES:
        return None

    if field_name.startswith("has_") or field_name == "is_available":
        return s.lower() in ("true", "1", "yes", "oui", "vrai")

    int_fields = {"country_id", "region_id", "pdv_id", "base_origin_id", "contract_id",
                  "capacity_eqp", "capacity_weight_kg", "eqp_count", "nb_colis", "dock_time_minutes",
                  "unload_time_per_eqp_minutes",
                  "sas_sec_capacity_eqc", "sas_frais_capacity_eqc", "sas_gel_capacity_eqc",
                  "origin_id", "destination_id",
                  "duration_minutes", "sequence_order"}
    if field_name in int_fields:
        try:
            return int(float(s))
        except ValueError:
            return s

    float_fields = {"latitude", "longitude", "fixed_cost", "cost_per_km", "cost_per_hour",
                    "fixed_daily_cost", "min_hours_per_day", "min_km_per_day", "consumption_coefficient",
                    "weight_kg", "distance_km", "total_km", "total_cost", "tax_per_km",
                    "sas_sec_surface_m2", "sas_frais_surface_m2", "sas_gel_surface_m2"}
    if field_name in float_fields:
        try:
            return float(s.replace(",", "."))
        except ValueError:
            return None

    # Nettoyer les préfixes d'enum Python (ex: "TemperatureClass.SEC" -> "SEC")
    # Strip Python enum class prefixes
    if "." in s and s.split(".", 1)[0].replace("_", "").isalpha():
        s = s.split(".", 1)[1]

    return s


async def _build_region_lookup(db: AsyncSession) -> dict[str, int]:
    """Construire un cache nom_région -> id / Build region name -> id lookup cache."""
    result = await db.execute(select(Region.id, Region.name))
    return {name.strip().lower(): rid for rid, name in result.all()}


def _resolve_region_id(value, region_lookup: dict[str, int]) -> int | None:
    """Résoudre region_id / Resolve region_id from numeric ID or region name."""
    if value is None:
        return None
    if isinstance(value, int):
        return value
    s = str(value).strip()
    try:
        return int(float(s))
    except (ValueError, TypeError):
        pass
    return region_lookup.get(s.lower())


UNIQUE_KEY_FIELDS = {
    "countries": "code",
    "regions": None,
    "bases": "code",
    "pdvs": "code",
    "suppliers": "code",
    "contracts": "code",
    "volumes": None,
    "distances": None,
    "km-tax": None,
}


async def _find_existing(db: AsyncSession, model_class, unique_field: str, value) -> object | None:
    """Chercher un enregistrement existant / Find existing record by unique key."""
    col = getattr(model_class, unique_field, None)
    if col is None:
        return None
    result = await db.execute(select(model_class).where(col == value))
    return result.scalar_one_or_none()


async def _build_code_lookup(db: AsyncSession) -> dict[str, tuple[str, int]]:
    """Construire un cache code -> (type, id) / Build code -> (type, db_id) lookup.
    Ajoute des variantes sans zéros initiaux pour gérer _coerce_value qui int() les codes.
    Adds zero-stripped variants to handle _coerce_value int() conversion of codes.
    """
    lookup: dict[str, tuple[str, int]] = {}
    for model, etype in [(PDV, "PDV"), (BaseLogistics, "BASE"), (Supplier, "SUPPLIER")]:
        result = await db.execute(select(model.id, model.code))
        for db_id, code in result.all():
            key = str(code).strip().lower()
            lookup[key] = (etype, db_id)
            # Variante sans zéros initiaux / Zero-stripped variant
            stripped = key.lstrip("0")
            if stripped and stripped not in lookup:
                lookup[stripped] = (etype, db_id)
    return lookup


def _resolve_pdv_type(value) -> str | None:
    """Résoudre le type de PDV via alias / Resolve PDV type via aliases."""
    if value is None:
        return None
    s = str(value).strip()
    if s.upper() in ("EXPRESS", "CONTACT", "SUPER_ALIMENTAIRE", "SUPER_GENERALISTE", "HYPER", "NETTO", "DRIVE", "URBAIN_PROXI"):
        return s.upper()
    return _PDV_TYPE_ALIASES.get(s.lower())


def _normalize_code(raw) -> str | None:
    """Normaliser un code matrice pour le lookup DB / Normalize matrix code for DB lookup.
    La matrice peut avoir 80, la DB peut avoir 080 → on compare sans zéros initiaux.
    """
    if raw is None:
        return None
    s = str(raw).strip()
    if not s or s.lower() == "x":
        return None
    try:
        return str(int(float(s)))
    except (ValueError, TypeError):
        return s.lower()


def _time_to_minutes(val) -> int | None:
    """Convertir datetime.time ou timedelta en minutes / Convert time value to minutes."""
    if val is None:
        return None
    if isinstance(val, dt_time):
        return val.hour * 60 + val.minute + round(val.second / 60)
    # openpyxl peut renvoyer un timedelta / openpyxl may return timedelta
    from datetime import timedelta
    if isinstance(val, timedelta):
        total = int(val.total_seconds())
        return total // 60
    # Valeur numérique brute (fraction de jour Excel) / Raw numeric value (Excel day fraction)
    if isinstance(val, (int, float)):
        total_minutes = round(val * 24 * 60)
        return total_minutes
    s = str(val).strip()
    if not s or s.lower() == "x":
        return None
    # Tenter HH:MM:SS ou HH:MM / Try HH:MM:SS or HH:MM
    parts = s.split(":")
    if len(parts) >= 2:
        try:
            h, m = int(parts[0]), int(parts[1])
            sec = int(parts[2]) if len(parts) > 2 else 0
            return h * 60 + m + round(sec / 60)
        except ValueError:
            pass
    return None


@router.post("/time-matrix")
async def import_time_matrix(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(require_permission("imports-exports", "create")),
):
    """
    Importer une matrice carrée de temps de trajet (Excel).
    Import a square travel-time matrix from Excel.

    Format attendu / Expected format:
    - Ligne 1 (header) : colonnes D+ = codes numériques des points
    - Ligne 2 : labels (ignorée)
    - Lignes 3+ : colonne A = code, colonnes D+ = temps (datetime.time)
    """
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Only .xlsx files are supported")

    content = await file.read()

    import openpyxl
    from io import BytesIO

    try:
        wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error reading Excel file: {e}")

    ws = wb.active
    if ws is None:
        raise HTTPException(status_code=400, detail="No active worksheet found")

    # 1. Lire les codes en header (ligne 1, colonnes D+) / Read header codes (row 1, cols D+)
    header_codes: list[str | None] = []
    for col_idx in range(4, ws.max_column + 1):  # D = col 4
        raw = ws.cell(row=1, column=col_idx).value
        header_codes.append(_normalize_code(raw))

    # 2. Construire le lookup code -> (type, db_id) / Build code lookup
    code_lookup = await _build_code_lookup(db)

    # Ajouter des variantes sans/avec zéros / Add variants without/with leading zeros
    extended_lookup: dict[str, tuple[str, int]] = {}
    for code_key, val in code_lookup.items():
        extended_lookup[code_key] = val
        # Sans zéros initiaux / Without leading zeros
        stripped = code_key.lstrip("0")
        if stripped and stripped not in extended_lookup:
            extended_lookup[stripped] = val
        # Avec zéros (3 chiffres) / With zeros (3 digits)
        try:
            padded = str(int(code_key)).zfill(3)
            if padded not in extended_lookup:
                extended_lookup[padded] = val
        except ValueError:
            pass

    def resolve_code(code: str | None) -> tuple[str, int] | None:
        """Résoudre un code matrice vers (type, id) / Resolve matrix code to (type, id)."""
        if code is None:
            return None
        for variant in [code, code.lstrip("0"), code.zfill(3)]:
            found = extended_lookup.get(variant)
            if found:
                return found
        return None

    # 3. Parser les données (lignes 3+) / Parse data rows (row 3+)
    created = 0
    updated = 0
    skipped = 0
    errors: list[str] = []

    for row_idx in range(3, ws.max_row + 1):
        row_code_raw = ws.cell(row=row_idx, column=1).value
        row_code = _normalize_code(row_code_raw)
        if row_code is None:
            continue

        row_resolved = resolve_code(row_code)
        if row_resolved is None:
            errors.append(f"Row {row_idx}: unknown code '{row_code_raw}' in column A")
            skipped += 1
            continue

        row_type, row_id = row_resolved

        for col_offset, col_code in enumerate(header_codes):
            if col_code is None:
                continue

            col_resolved = resolve_code(col_code)
            if col_resolved is None:
                continue

            col_type, col_id = col_resolved

            # Ignorer la diagonale / Skip diagonal
            if row_type == col_type and row_id == col_id:
                continue

            cell_val = ws.cell(row=row_idx, column=4 + col_offset).value
            minutes = _time_to_minutes(cell_val)
            if minutes is None:
                continue

            # Chercher l'entrée existante (bidirectionnelle) / Find existing entry (bidirectional)
            result = await db.execute(
                select(DistanceMatrix).where(
                    DistanceMatrix.origin_type == row_type,
                    DistanceMatrix.origin_id == row_id,
                    DistanceMatrix.destination_type == col_type,
                    DistanceMatrix.destination_id == col_id,
                )
            )
            existing = result.scalar_one_or_none()

            if not existing:
                result = await db.execute(
                    select(DistanceMatrix).where(
                        DistanceMatrix.origin_type == col_type,
                        DistanceMatrix.origin_id == col_id,
                        DistanceMatrix.destination_type == row_type,
                        DistanceMatrix.destination_id == row_id,
                    )
                )
                existing = result.scalar_one_or_none()

            if existing:
                if existing.duration_minutes != minutes:
                    existing.duration_minutes = minutes
                    updated += 1
                else:
                    skipped += 1
            else:
                db.add(DistanceMatrix(
                    origin_type=row_type,
                    origin_id=row_id,
                    destination_type=col_type,
                    destination_id=col_id,
                    distance_km=0,
                    duration_minutes=minutes,
                ))
                created += 1

    if created > 0 or updated > 0:
        try:
            await db.flush()
        except Exception as e:
            await db.rollback()
            raise HTTPException(status_code=400, detail=f"Database error: {e}")

    return {
        "status": "success",
        "created": created,
        "updated": updated,
        "skipped": skipped,
        "errors": errors[:20],
        "message": f"{created} created, {updated} updated, {skipped} skipped",
    }


@router.post("/{entity_type}")
async def import_data(
    entity_type: str,
    file: UploadFile = File(...),
    mode: str = Query("check", description="check (default) = warn on duplicates, replace = delete existing, append = skip check"),
    dispatch_date: str | None = Query(None, description="Date de répartition (YYYY-MM-DD) — volumes only"),
    dispatch_time: str | None = Query(None, description="Heure de répartition (HH:MM) — volumes only"),
    activity_type: str | None = Query(None, description="Activité : SUIVI ou MEAV — volumes only"),
    promo_start_date: str | None = Query(None, description="Date début promo (YYYY-MM-DD) — MEAV only"),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(require_permission("imports-exports", "create")),
):
    """
    Importer des données depuis un fichier CSV ou Excel.
    Import data from a CSV or Excel file.
    mode: check (détecte doublons), replace (remplace existants), append (ajoute sans vérifier)
    """
    if entity_type not in ENTITY_MODEL_MAP:
        raise HTTPException(status_code=400, detail=f"Invalid entity type. Allowed: {list(ENTITY_MODEL_MAP.keys())}")

    if not file.filename:
        raise HTTPException(status_code=400, detail="No file provided")

    ext = file.filename.rsplit(".", 1)[-1].lower()
    if ext not in ("csv", "xlsx", "xls"):
        raise HTTPException(status_code=400, detail="Only CSV and Excel files are supported")

    content = await file.read()
    try:
        rows = ImportService.parse_file(content, file.filename)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error parsing file: {e}")

    if not rows:
        raise HTTPException(status_code=400, detail="No data found in file")

    region_lookup = await _build_region_lookup(db)

    # Lookups typés par entité (code -> db_id) avec variantes sans zéros
    # Typed lookups per entity (code -> db_id) with zero-stripped variants
    typed_lookups: dict[str, dict[str, int]] = {}
    if entity_type in ("distances", "km-tax", "volumes"):
        for model, etype in [(PDV, "PDV"), (BaseLogistics, "BASE"), (Supplier, "SUPPLIER")]:
            lk: dict[str, int] = {}
            result = await db.execute(select(model.id, model.code))
            for db_id, code in result.all():
                key = str(code).strip().lower()
                lk[key] = db_id
                stripped = key.lstrip("0")
                if stripped and stripped not in lk:
                    lk[stripped] = db_id
            typed_lookups[etype] = lk

    # --- Détection doublons volumes / Volume duplicate detection ---
    if entity_type == "volumes" and mode != "append":
        base_ids_in_file: set[int] = set()
        base_lk = typed_lookups.get("BASE", {})
        for row in rows:
            for key, val in row.items():
                ck = key.strip().lower().replace(" ", "_")
                if ck == "base_origin_id" and val is not None:
                    code_str = str(val).strip().lower()
                    resolved = base_lk.get(code_str)
                    if not resolved:
                        try:
                            resolved = base_lk.get(str(int(float(code_str))))
                        except (ValueError, TypeError):
                            pass
                    if resolved:
                        base_ids_in_file.add(resolved)

        if base_ids_in_file:
            # Si dispatch_date fourni → détection par (dispatch_date, dispatch_time, base_origin_id)
            # If dispatch_date provided → detect by (dispatch_date, dispatch_time, base_origin_id)
            if dispatch_date:
                dup_query = select(Volume.dispatch_date, Volume.dispatch_time, Volume.base_origin_id, func.count(Volume.id)).where(
                    Volume.dispatch_date == dispatch_date,
                    Volume.base_origin_id.in_(list(base_ids_in_file)),
                )
                if dispatch_time:
                    dup_query = dup_query.where(Volume.dispatch_time == dispatch_time)
                dup_query = dup_query.group_by(Volume.dispatch_date, Volume.dispatch_time, Volume.base_origin_id)
                result = await db.execute(dup_query)
                existing_groups = result.all()
            else:
                # Fallback : détection par (date, base_origin_id) / Fallback: detect by (date, base_origin_id)
                dates_in_file: set[str] = set()
                for row in rows:
                    for key, val in row.items():
                        ck = key.strip().lower().replace(" ", "_")
                        if ck == "date" and val is not None:
                            coerced = _coerce_value(val, "date")
                            if coerced:
                                dates_in_file.add(str(coerced))
                if dates_in_file:
                    result = await db.execute(
                        select(Volume.date, Volume.base_origin_id, func.count(Volume.id))
                        .where(Volume.date.in_(list(dates_in_file)), Volume.base_origin_id.in_(list(base_ids_in_file)))
                        .group_by(Volume.date, Volume.base_origin_id)
                    )
                    existing_groups = result.all()
                else:
                    existing_groups = []

            if existing_groups:
                if mode == "check":
                    base_result = await db.execute(
                        select(BaseLogistics.id, BaseLogistics.code, BaseLogistics.name)
                        .where(BaseLogistics.id.in_(list(base_ids_in_file)))
                    )
                    base_names = {bid: f"{code} — {name}" for bid, code, name in base_result.all()}
                    if dispatch_date:
                        existing_details = [
                            {"dispatch_date": dd or "", "dispatch_time": dt or "", "base": base_names.get(b, str(b)), "count": c}
                            for dd, dt, b, c in existing_groups
                        ]
                    else:
                        existing_details = [
                            {"date": d, "base": base_names.get(b, str(b)), "count": c}
                            for d, b, c in existing_groups
                        ]
                    total = sum(e["count"] for e in existing_details)
                    return {
                        "status": "duplicate_warning",
                        "existing": existing_details,
                        "total_existing": total,
                        "new_row_count": len(rows),
                        "message": f"{total} volumes existent déjà pour les mêmes dates/bases",
                    }
                elif mode == "replace":
                    if dispatch_date:
                        del_query = sa_delete(Volume).where(
                            Volume.dispatch_date == dispatch_date,
                            Volume.base_origin_id.in_(list(base_ids_in_file)),
                        )
                        if dispatch_time:
                            del_query = del_query.where(Volume.dispatch_time == dispatch_time)
                        await db.execute(del_query)
                    else:
                        dates_in_file_list = list(dates_in_file) if dates_in_file else []
                        if dates_in_file_list:
                            await db.execute(
                                sa_delete(Volume).where(
                                    Volume.date.in_(dates_in_file_list),
                                    Volume.base_origin_id.in_(list(base_ids_in_file)),
                                )
                            )

    model_class = ENTITY_MODEL_MAP[entity_type]
    allowed_fields = set(ImportService.ENTITY_FIELDS.get(entity_type, []))
    required = REQUIRED_FIELDS.get(entity_type, set())
    unique_field = UNIQUE_KEY_FIELDS.get(entity_type)
    created = 0
    updated = 0
    skipped = 0
    errors = []

    for i, row in enumerate(rows):
        try:
            data = {}
            for key, val in row.items():
                clean_key = key.strip().lower().replace(" ", "_")
                if clean_key in allowed_fields:
                    data[clean_key] = _coerce_value(val, clean_key)

            if not data:
                skipped += 1
                continue

            if "region_id" in data and data["region_id"] is not None:
                if isinstance(data["region_id"], str):
                    resolved = _resolve_region_id(data["region_id"], region_lookup)
                    if resolved is None:
                        errors.append(f"Row {i + 2}: region '{data['region_id']}' not found in database")
                        skipped += 1
                        continue
                    data["region_id"] = resolved

            if entity_type == "pdvs" and "type" in data and data["type"] is not None:
                resolved_type = _resolve_pdv_type(data["type"])
                if resolved_type is None:
                    errors.append(f"Row {i + 2}: unknown PDV type '{data['type']}'")
                    skipped += 1
                    continue
                data["type"] = resolved_type

            if "code" in data and data["code"] is not None:
                data["code"] = str(data["code"]).strip()

            if entity_type in ("distances", "km-tax") and typed_lookups:
                for prefix in ("origin", "destination"):
                    id_key = f"{prefix}_id"
                    type_key = f"{prefix}_type"
                    raw_id = data.get(id_key)
                    if raw_id is not None:
                        code_str = str(raw_id).strip().lower()
                        # Si le type est fourni dans le CSV, utiliser le lookup typé
                        # If type is provided in CSV, use the typed lookup
                        csv_type = str(data.get(type_key, "")).strip().upper()
                        if csv_type and csv_type in typed_lookups:
                            resolved_id = typed_lookups[csv_type].get(code_str)
                            if resolved_id is not None:
                                data[type_key] = csv_type
                                data[id_key] = resolved_id
                            else:
                                errors.append(f"Row {i + 2}: unknown {csv_type} code '{raw_id}' for {id_key}")
                                skipped += 1
                                data = {}
                                break
                        else:
                            # Pas de type → chercher dans tous les lookups
                            # No type → search all lookups
                            found = False
                            for etype, lk in typed_lookups.items():
                                resolved_id = lk.get(code_str)
                                if resolved_id is not None:
                                    data[type_key] = etype
                                    data[id_key] = resolved_id
                                    found = True
                                    break
                            if not found:
                                errors.append(f"Row {i + 2}: unknown code '{raw_id}' for {id_key}")
                                skipped += 1
                                data = {}
                                break
                if not data:
                    continue

            # Volumes : résoudre pdv_id et base_origin_id par code (lookups typés)
            # Volumes: resolve pdv_id and base_origin_id from code (typed lookups)
            if entity_type == "volumes" and typed_lookups:
                fk_map = {"pdv_id": "PDV", "base_origin_id": "BASE"}
                for fk_field, fk_type in fk_map.items():
                    raw_val = data.get(fk_field)
                    if raw_val is not None:
                        code_str = str(raw_val).strip().lower()
                        resolved_id = typed_lookups[fk_type].get(code_str)
                        if resolved_id is not None:
                            data[fk_field] = resolved_id
                        else:
                            errors.append(f"Row {i + 2}: unknown {fk_type} code '{raw_val}' for {fk_field}")
                            skipped += 1
                            data = {}
                            break
                if not data:
                    continue

            # Injecter dispatch_date/dispatch_time dans chaque volume importé
            # Inject dispatch metadata into each imported volume
            if entity_type == "volumes":
                if dispatch_date:
                    data["dispatch_date"] = dispatch_date
                if dispatch_time:
                    data["dispatch_time"] = dispatch_time
                if activity_type:
                    data["activity_type"] = activity_type
                if promo_start_date:
                    data["promo_start_date"] = promo_start_date

            missing = [f for f in required if f not in data or data[f] is None]
            if missing:
                errors.append(f"Row {i + 2}: missing required fields: {', '.join(missing)}")
                skipped += 1
                continue

            existing = None
            if unique_field and unique_field in data and data[unique_field] is not None:
                existing = await _find_existing(db, model_class, unique_field, data[unique_field])
            elif entity_type == "distances":
                result = await db.execute(
                    select(DistanceMatrix).where(
                        DistanceMatrix.origin_type == data.get("origin_type"),
                        DistanceMatrix.origin_id == data.get("origin_id"),
                        DistanceMatrix.destination_type == data.get("destination_type"),
                        DistanceMatrix.destination_id == data.get("destination_id"),
                    )
                )
                existing = result.scalar_one_or_none()
            elif entity_type == "km-tax":
                result = await db.execute(
                    select(KmTax).where(
                        KmTax.origin_type == data.get("origin_type"),
                        KmTax.origin_id == data.get("origin_id"),
                        KmTax.destination_type == data.get("destination_type"),
                        KmTax.destination_id == data.get("destination_id"),
                    )
                )
                existing = result.scalar_one_or_none()

            if existing:
                for k, v in data.items():
                    if k != unique_field:
                        setattr(existing, k, v)
                updated += 1
            else:
                obj = model_class(**data)
                db.add(obj)
                created += 1
        except Exception as e:
            errors.append(f"Row {i + 2}: {e}")
            skipped += 1

    if created > 0 or updated > 0:
        try:
            await db.flush()
        except Exception as e:
            await db.rollback()
            raise HTTPException(status_code=400, detail=f"Database error during import: {e}")

    return {
        "status": "success",
        "created": created,
        "updated": updated,
        "skipped": skipped,
        "total_rows": len(rows),
        "errors": errors[:20],
        "message": f"{created} created, {updated} updated, {skipped} skipped (out of {len(rows)} rows)",
    }
