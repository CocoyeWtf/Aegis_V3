"""Tests export WMS Infolog (TMS_vers_wms) / WMS Infolog export tests.

Vérifie le format attendu par la macro d'encodage Infolog :
- une ligne par arrêt PDV ;
- tours rangés par priorité ERT ;
- PDV de chaque tour en ordre INVERSE (dernier livré encodé en premier) ;
- index global décroissant par tour, bloc contigu par tour ;
- code transporteur issu du paramètre wms_infolog_carrier_code.
"""

import io
import uuid

import pytest
from openpyxl import load_workbook


async def _make_base(db_session, region):
    from app.models.base_logistics import BaseLogistics

    base = BaseLogistics(
        code=f"B{uuid.uuid4().hex[:5].upper()}",
        name="Base WMS Test",
        region_id=region.id,
    )
    db_session.add(base)
    await db_session.commit()
    await db_session.refresh(base)
    return base


async def _make_pdv(db_session, region, code):
    from app.models.pdv import PDV, PDVType

    pdv = PDV(
        code=code,
        name=f"PDV {code}",
        type=PDVType.HYPER,
        region_id=region.id,
    )
    db_session.add(pdv)
    await db_session.commit()
    await db_session.refresh(pdv)
    return pdv


async def _make_tour(db_session, base, *, code, date, priority, departure, driver_code, pdv_codes_in_delivery_order, region):
    """Crée un tour planifié avec ses arrêts dans l'ordre de livraison 1..n."""
    from app.models.tour import Tour, TourStatus
    from app.models.tour_stop import TourStop

    tour = Tour(
        date=date,
        code=code,
        base_id=base.id,
        status=TourStatus.VALIDATED,
        departure_time=departure,
        delivery_date=date,
        priority=priority,
        driver_code_infolog=driver_code,
    )
    db_session.add(tour)
    await db_session.commit()
    await db_session.refresh(tour)

    for seq, pcode in enumerate(pdv_codes_in_delivery_order, start=1):
        pdv = await _make_pdv(db_session, region, pcode)
        db_session.add(TourStop(
            tour_id=tour.id,
            pdv_id=pdv.id,
            sequence_order=seq,
            eqp_count=1,
        ))
    await db_session.commit()
    return tour


@pytest.mark.asyncio
async def test_wms_export_reversal_and_index(client, db_session, test_region):
    base = await _make_base(db_session, test_region)
    date = "2026-06-09"

    # Tour 1 (priorité 1) : 3 PDV livrés A→B→C / delivered A then B then C
    await _make_tour(
        db_session, base, code=f"WMS1-{uuid.uuid4().hex[:4]}", date=date, priority=1,
        departure="08:00", driver_code="08000123",
        pdv_codes_in_delivery_order=["05198", "09918", "07893"], region=test_region,
    )
    # Tour 2 (priorité 2) : 2 PDV livrés D→E / delivered D then E
    await _make_tour(
        db_session, base, code=f"WMS2-{uuid.uuid4().hex[:4]}", date=date, priority=2,
        departure="09:00", driver_code="08000124",
        pdv_codes_in_delivery_order=["02364", "06912"], region=test_region,
    )

    resp = await client.get(f"/api/exports/wms-infolog", params={"date": date, "base_id": base.id})
    assert resp.status_code == 200, resp.text

    wb = load_workbook(io.BytesIO(resp.content))
    ws = wb["Export"]
    rows = [tuple(r) for r in ws.iter_rows(values_only=True)]

    # 5 arrêts au total / 5 stops total
    assert len(rows) == 5

    # Colonnes A,B,C,G pertinentes : (ordre, pdv, code_chauffeur, ..., index)
    # Tour 1 inversé : C,B,A en B(pdv) ; index global 3,2,1
    assert [r[1] for r in rows[:3]] == ["07893", "09918", "05198"]  # ordre inverse
    assert [r[0] for r in rows[:3]] == [1, 1, 1]                    # ordre ERT
    assert [r[2] for r in rows[:3]] == ["08000123"] * 3            # code chauffeur
    assert [r[6] for r in rows[:3]] == [3, 2, 1]                    # index décroissant

    # Tour 2 inversé : E,D ; index global contigu 5,4
    assert [r[1] for r in rows[3:]] == ["06912", "02364"]
    assert [r[0] for r in rows[3:]] == [2, 2]
    assert [r[6] for r in rows[3:]] == [5, 4]

    # Heure de départ en texte HH:MM:SS (col F et H identiques)
    assert rows[0][5] == "08:00:00"
    assert rows[0][7] == "08:00:00"

    # Code transporteur par défaut (aucun paramètre défini)
    assert all(r[3] == "08000888" for r in rows)


@pytest.mark.asyncio
async def test_wms_export_carrier_code_parameter(client, db_session, test_region):
    from app.models.parameter import Parameter

    base = await _make_base(db_session, test_region)
    date = "2026-06-10"

    db_session.add(Parameter(
        key="wms_infolog_carrier_code", value="08009999", value_type="string", region_id=None,
    ))
    await db_session.commit()

    await _make_tour(
        db_session, base, code=f"WMS3-{uuid.uuid4().hex[:4]}", date=date, priority=1,
        departure="07:30", driver_code="08000200",
        pdv_codes_in_delivery_order=["11111", "22222"], region=test_region,
    )

    resp = await client.get(f"/api/exports/wms-infolog", params={"date": date, "base_id": base.id})
    assert resp.status_code == 200, resp.text
    wb = load_workbook(io.BytesIO(resp.content))
    rows = [tuple(r) for r in wb["Export"].iter_rows(values_only=True)]
    assert all(r[3] == "08009999" for r in rows)
